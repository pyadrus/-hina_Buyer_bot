import os
import sqlite3

import openpyxl
from aiogram import types, F
from aiogram.filters import Command
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import FSInputFile
from aiogram.types import Message
from loguru import logger

from system.dispatcher import ADMIN_USER_ID
from system.dispatcher import bot
from system.dispatcher import dp
from system.dispatcher import router


@router.message(Command("help"))
async def help_handler(message: Message, state: FSMContext):
    """Админ панель"""
    if message.from_user.id == ADMIN_USER_ID:
        await message.answer("Команды админа:\n\n"
                             "/edit_services_and_prices - редактирование: ⭐️ Услуги и цены\n"
                             "/edit - редактирование: пост приветствиt\n"
                             "/edit_self_purchase - редактирование: 🛍 Самовыкуп\n"
                             "/edit_product_search - редактирование: Подбор товара\n"
                             "/edit_search_in_china - редактирование: Поиск поставщика в Китае\n"
                             "/edit_warranty_service - редактирование: Выкуп товаров\n"
                             "/edit_delivery_in_china - редактирование: Доставка в Китае\n"
                             "/edit_payment_process - редактирование: Как совершается оплата?\n"
                             "/edit_order_form - редактирование: 🗒 Бланк заказа\n"
                             "/edit_payment_options - редактирование: Какие платежи меня ожидают?\n"
                             "/edit_reviews - редактирование: 💌 Отзывы\n"
                             "/edit_bag_tape - редактирование: Мешок + скотч\n"
                             "/edit_box_bag_tape - редактирование: Коробка + мешок + скотч\n"
                             "/edit_cardboard_corners_bag_tape - редактирование: Картонные уголки + мешок + скотч\n"
                             "/edit_pallet_crate - редактирование: Паллет в обрешетке\n"
                             "edit_pallet_with_solid_wooden_box - редактирование: Паллет с глухим деревянным коробом\n"
                             "/edit_types_packaging_handlers - редактирование: Назад к видам упаковки\n"
                             "/edit_wooden_sheathing_bag_tape - редактирование: Деревянная обрешетка + мешок + скотч\n\n"
                             "/start - начальное меню\n")
    else:
        await message.reply("У вас нет прав на выполнение этой команды.")


# Функция для создания файла Excel с данными заказов
def create_excel_file(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'Имя'
    sheet['C1'] = 'Фамилия'
    sheet['D1'] = 'Город'
    sheet['E1'] = 'Номер телефона'
    sheet['F1'] = 'Дата регистрации'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # Имя
        sheet.cell(row=index, column=3).value = order[2]  # Фамилия
        sheet.cell(row=index, column=4).value = order[3]  # Город
        sheet.cell(row=index, column=5).value = order[4]  # Номер телефона
        sheet.cell(row=index, column=6).value = order[5]  # Дата регистрации

    return workbook


@router.message(Command("get_a_list_of_users_registered_in_the_bot"))
async def export_data(message: types.Message, state: FSMContext):
    """Получение списка зарегистрированных пользователей"""
    await state.clear()  # Очищаем состояние
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        # Подключение к базе данных SQLite
        conn = sqlite3.connect('your_database.db')
        cursor = conn.cursor()
        # Получение данных из базы данных
        cursor.execute("SELECT * FROM users")
        orders = cursor.fetchall()
        # Создание файла Excel
        workbook = create_excel_file(orders)
        filename = 'Зарегистрированные пользователи в боте.xlsx'
        workbook.save(filename)  # Сохранение файла
        text = ("Данные пользователей зарегистрированных в боте\n\n"
                "Для возврата в начальное меню нажми на /start или /help")
        file = FSInputFile(filename)
        await bot.send_document(message.from_user.id, document=file, caption=text)  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


def create_excel_file_start(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'username'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Дата запуска бота'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # username
        sheet.cell(row=index, column=3).value = order[2]  # Имя
        sheet.cell(row=index, column=4).value = order[3]  # Фамилия
        sheet.cell(row=index, column=5).value = order[4]  # Дата запуска бота

    return workbook


@router.message(Command("get_users_who_launched_the_bot"))
async def get_users_who_launched_the_bot(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.clear()  # Очищаем состояние
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        # Подключение к базе данных SQLite
        conn = sqlite3.connect('your_database.db')
        cursor = conn.cursor()
        # Получение данных из базы данных
        cursor.execute("SELECT * FROM users_start")
        orders = cursor.fetchall()
        # Создание файла Excel
        workbook = create_excel_file_start(orders)
        filename = 'Данные пользователей запустивших бота.xlsx'
        workbook.save(filename)  # Сохранение файла
        # file = InputFile(filename)
        file = FSInputFile(filename)
        text = ("Данные пользователей зарегистрированных в боте\n\n"
                "Для возврата в начальное меню нажми на /start или /help")
        await bot.send_document(message.from_user.id, document=file, caption=text)  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


class MyStates(StatesGroup):
    waiting_for_message = State()
    waiting_for_image = State()
    waiting_for_caption = State()


@router.message(Command("send_an_image_to_bot_users"))
async def send_an_image_to_bot_users(message: types.Message, state: FSMContext):
    """Запрашивает изображение у администратора"""
    await state.clear()  # Очищаем состояние
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        await bot.send_message(message.from_user.id, text="Загрузите изображение для рассылки:")
        await state.set_state(MyStates.waiting_for_image)
    except Exception as e:
        logger.error(e)


@router.message(StateFilter(MyStates.waiting_for_image))
async def process_send_image(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать загруженного изображения и переходить в состояние "ожидание подписи"
    """

    await state.update_data(photo=message.photo[-1].file_id)
    await bot.send_message(message.from_user.id, text="Введите подпись к изображению:")
    await state.set_state(MyStates.waiting_for_caption)


@router.message(StateFilter(MyStates.waiting_for_caption))
async def process_send_image_with_caption(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать введенной подписи и выполнять рассылку
    """
    state_data = await state.get_data()  # Retrieve state data
    # async with state.proxy() as data:
    #     data['caption'] = message.text
    # Store the caption in state data
    state_data['caption'] = message.text
    # Get the photo and caption from state data
    photo = state_data.get('photo')
    caption = state_data.get('caption')
    # Получаем список уникальных ID пользователей из базы данных
    user_ids = get_user_ids()
    if user_ids:
        # Рассылка изображения с подписью всем пользователям из списка
        for user_id in user_ids:
            try:
                # Отправляем изображение с подписью
                await bot.send_photo(user_id, photo, caption=caption)
            except Exception as e:
                print(f"Ошибка при отправке изображения с подписью пользователю {user_id}: {str(e)}")
    await message.answer("Изображение успешно разослано всем пользователям.")
    await state.clear()


@router.message(Command("send_a_message_to_bot_users"))
async def send_a_message_to_bot_users(message: types.Message, state: FSMContext):
    """Запрашивает текст сообщения у администратора"""
    await state.clear()
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        await bot.send_message(message.from_user.id, text="Введите текст для рассылки:")
        await state.set_state(MyStates.waiting_for_message)
    except Exception as e:
        logger.error(e)


@router.message(StateFilter(MyStates.waiting_for_message), F.TEXT)
# @dp.message_handler(state=MyStates.waiting_for_message, content_types=types.ContentType.TEXT)
async def process_send_message(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать введенного текста и выполнять рассылку
    """
    async with state.proxy() as data:
        data['message_text'] = message.text
    # Получаем список уникальных ID пользователей из базы данных
    user_ids = get_user_ids()
    if user_ids:
        # Рассылка сообщения всем пользователям из списка
        for user_id in user_ids:
            try:
                await bot.send_message(user_id, data['message_text'], parse_mode=ParseMode.MARKDOWN)
            except Exception as e:
                print(f"Ошибка при отправке сообщения пользователю {user_id}: {str(e)}")
    await message.answer("Сообщение успешно разослано всем пользователям.")
    await state.clear()


def get_user_ids():
    """Получаем уникальные ID пользователей из базы данных"""
    try:
        conn = sqlite3.connect('your_database.db')  # Замените 'your_database.db' на имя вашей базы данных
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT user_id FROM users_start")
        user_ids = [row[0] for row in cursor.fetchall()]
        return user_ids
    except Exception as e:
        print(f"Ошибка при получении ID пользователей из базы данных: {str(e)}")
        return []


def register_handlers_admin():
    dp.message.register(help_handler)
    dp.message.register(export_data)
    dp.message.register(get_users_who_launched_the_bot)
    dp.message.register(send_a_message_to_bot_users)
    dp.message.register(send_an_image_to_bot_users)
